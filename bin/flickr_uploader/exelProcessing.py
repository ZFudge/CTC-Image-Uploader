#
# Process the excel sheets for youtube videos and github code
# Retrive short url.
#
# type codes:
# y: youtube video
# g: github code
#
# state codes:
# 0: not shortened
# 1: shortened
#
import openpyxl
from mainDB import CTCPhotoDB
from CreateShortLinks import getShortURL

videoSheetLocation="data/CSV ctc vid.xlsx"
codeSheetLocation="data/Github bitly sheet.xlsx"

photoDB=CTCPhotoDB()


#
# Util function for getting the type name of an
# extra record
#
#
def getFullType(typeCode):
	return { \
		"y":"Youtube video", \
		"g":"Github code" \
	}[typeCode]

#
# Util function for getting the in-set order 
# by reading the last digits of a shortCode
#
#
def getOrderInSet(shortCode):
	identifier=shortCode.split("-")[-2:]
	if identifier[0].isdigit():
		return int(identifier[1])-1
	else:
		return 0



#
# get the active sheet of a openpyxl object from a file 
# location string
#
#
def getSheet(sheetLocation):
	wb = openpyxl.load_workbook(sheetLocation)
	sheet=wb.active
	return sheet

#
# Read data from the exel sheet of youtube videos and 
# save it in the database
#
#
def processVideoSheet():
	sheet=getSheet(videoSheetLocation)
	for row in range(1,sheet.max_row):
		name=sheet.cell(row=row,column=1).value
		link=sheet.cell(row=row,column=2).value
		shortCode=sheet.cell(row=row,column=3).value
		if shortCode!=None and link!=None:
			orderInSet=getOrderInSet(shortCode)
			shortCode="ctc-y-"+shortCode
			photoDB.addExtra({"name":name, "short_code":shortCode, "hosted_url":link, "type":"y", "order_in_set":orderInSet})
			print name, shortCode
	photoDB.commit()


#
# Read data from the exel sheet of github codes and 
# save it in the database
#
#
def processCodeSheet():
	sheet=getSheet(codeSheetLocation)
	for row in range(4,sheet.max_row):
		name=sheet.cell(row=row,column=1).value
		link=sheet.cell(row=row,column=3).value
		shortCode=sheet.cell(row=row,column=5).value
		if shortCode!=None and link!=None:
			orderInSet=getOrderInSet(shortCode)
			shortCode="ctc-g-"+shortCode
			photoDB.addExtra({"name":name, "short_code":shortCode, "hosted_url":link, "type":"g", "order_in_set":orderInSet})
			print name, shortCode
	photoDB.commit()


#
# Request short URL for extras, and save the progress as state
#
#
def getShortURLForExtras():
	cmd="""
	SELECT * FROM extras
	WHERE state == 0
	"""
	toGet=photoDB.makeQuery(cmd)[0].fetchall()

	for one in toGet[0:1]:
		#print one["hosted_url"]
		getShortURL({ \
			"hosted_url":one["hosted_url"], \
			"keyword":one["short_code"], \
			"title":getFullType(one["type"])+" "+one["name"] \
			})
		photoDB.modifyExtraState(one["short_code"],1)